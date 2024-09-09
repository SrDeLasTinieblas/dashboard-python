# dashboard_utils.py
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime, time
from io import BytesIO

def get_cookie(key):
    return st.session_state.get(key, None)

def generar_excel(alumnas_df, usuario, logo_path):
    wb = Workbook()
    ws = wb.active
    
    # Título del sistema y datos de descarga
    ws["A1"] = "Sistema de Asistencias - Taller de Pole Dance"
    ws["A2"] = f"Generado por: {usuario}"
    ws["A3"] = f"Fecha y Hora: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Agregar logo
    img = Image(logo_path)  # Especificar la ruta de la imagen del logo
    img.height = 80  # Ajustar la altura del logo
    img.width = 150  # Ajustar el ancho del logo
    ws.add_image(img, "D1")  # Colocar la imagen en la celda D1
    
    # Título de la tabla
    ws["A5"] = "Listado de Alumnas"
    
    # Añadir los datos de las alumnas
    for r_idx, row in enumerate(dataframe_to_rows(alumnas_df, index=False, header=True), 6):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Ajustar el tamaño de las columnas
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    # Guardar el archivo en un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def mostrar_dashboard():
    # Datos ficticios de ejemplo para las alumnas
    alumnas_data = {
        'Nombre': ['Ana García', 'Sofía Torres', 'Lucía Sánchez', 'Valentina Díaz', 'María López'],
        'Clases Restantes': [5, 8, 3, 10, 7],
        'Asistencias Totales': [15, 22, 17, 10, 18]
    }

    # Convertir los datos en un DataFrame
    alumnas_df = pd.DataFrame(alumnas_data)

    st.title("Sistema de Asistencias - Taller de Pole Dance")

    # Texto de bienvenida
    st.write("Bienvenido/a al sistema de gestión de asistencias del taller de pole dance. Aquí puedes registrar asistencias, ver el historial y gestionar a las alumnas.")

    # Sección 1: Mostrar lista de alumnas y clases restantes
    st.subheader("Lista de Alumnas y Clases Restantes")
    st.write(alumnas_df)

    # Sección 2: Registrar nueva asistencia
    st.subheader("Registrar Asistencia")
    seleccion_alumna = st.selectbox("Selecciona una alumna para registrar la asistencia", alumnas_df['Nombre'])

    # Botón para registrar asistencia
    if st.button("Registrar Asistencia"):
        # Encontrar a la alumna seleccionada y restar una clase
        index = alumnas_df[alumnas_df['Nombre'] == seleccion_alumna].index[0]
        if alumnas_df.at[index, 'Clases Restantes'] > 0:
            alumnas_df.at[index, 'Clases Restantes'] -= 1
            alumnas_df.at[index, 'Asistencias Totales'] += 1
            st.success(f"Asistencia registrada para {seleccion_alumna}")
        else:
            st.error(f"{seleccion_alumna} no tiene clases restantes. Debe renovar su plan.")

    # Mostrar datos actualizados
    st.write("Datos actualizados:")
    st.write(alumnas_df)

    # Sección 3: Historial de asistencia
    st.subheader("Historial de Asistencia (Últimas 5 Clases)")
    historial_data = {
        'Fecha': ['2024-09-01', '2024-08-28', '2024-08-25', '2024-08-20', '2024-08-15'],
        'Alumna': [seleccion_alumna] * 5,  # Historial ficticio de la alumna seleccionada
        'Clase Número': [15, 14, 13, 12, 11]
    }

    historial_df = pd.DataFrame(historial_data)
    st.write(historial_df)

    # Sección 4: Gráfico de asistencias totales
    st.subheader("Gráfico de Asistencias Totales")
    st.bar_chart(alumnas_df.set_index('Nombre')['Asistencias Totales'])

    # Sección 5: Gestión de alertas
    st.subheader("Alertas de Clases Restantes")
    for index, row in alumnas_df.iterrows():
        if row['Clases Restantes'] <= 2:
            st.warning(f"{row['Nombre']} tiene solo {row['Clases Restantes']} clase(s) restante(s).")

    logo_path = "Logo_PoleDance.jpg"  # Ruta del logo de la empresa (asegúrate de tener la imagen en tu directorio)
    excel_file = generar_excel(alumnas_df, 'ADMIN', logo_path)

    st.download_button(
        label="Descargar Excel",
        data=excel_file,
        file_name="alumnas_asistencia.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
